#!/usr/bin/env python3
"""Build the editable Raijin AWS-to-OCI migration cost calculator workbook."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT = Path(__file__).resolve().parents[1] / "docs" / "raijin-cost-calculator.xlsx"
NAVY = "101A2D"
PANEL = "19263E"
BLUE = "3B82F6"
TEAL = "14B8A6"
ORANGE = "E87D18"
GREEN = "166534"
YELLOW = "B7791F"
WHITE = "F8FAFC"
MUTED = "C5D2E8"
INPUT = "FFF2CC"
THIN = Side(style="thin", color="4B5E7A")


def title(ws, value, subtitle=None):
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = value
    cell.font = Font(size=18, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 32
    if subtitle:
        ws.merge_cells("A2:H2")
        cell = ws["A2"]
        cell.value = subtitle
        cell.font = Font(italic=True, color=MUTED)
        cell.fill = PatternFill("solid", fgColor=PANEL)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[2].height = 34
    ws.sheet_view.showGridLines = False


def heading(cell):
    cell.font = Font(bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=BLUE)
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    cell.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def body(cell, input_cell=False):
    cell.fill = PatternFill("solid", fgColor=INPUT if input_cell else PANEL)
    cell.font = Font(color="111827" if input_cell else WHITE)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def table(ws, row, headers, widths=None):
    for col, name in enumerate(headers, 1):
        c = ws.cell(row, col, name)
        heading(c)
        if widths:
            ws.column_dimensions[c.column_letter].width = widths[col - 1]


def currency(cell):
    cell.number_format = 'US$ #,##0.00;[Red]-US$ #,##0.00;—'


def number(cell, places=2):
    cell.number_format = f'#,##0.{"0" * places};[Red]-#,##0.{"0" * places};—'


def build_readme(wb):
    ws = wb.active
    ws.title = "Read me"
    title(ws, "RAIJIN — Migration Cost Calculator", "Editable estimate for AWS S3 / Glacier Deep Archive to OCI Object Storage. Yellow cells are inputs; formulas are protected only by convention, not by password.")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 110
    items = [
        ("1. Tariffs", "Maintain public or contracted rates by region in ‘Tariffs by region’. The sa-east-1 starter values were collected by Raijin on 2026-08-19; validate them before approving spend."),
        ("Refresh AWS public rates", "Run: python3 tools/refresh_cost_calculator_rates.py --regions sa-east-1,us-east-1. It calls the same public Amazon S3 and AWS Data Transfer catalog endpoints used by Raijin. It updates AWS public rate columns only; FastConnect/Direct Connect, Deep Archive storage and OCI fields remain under customer control."),
        ("2. Main inputs", "Select the AWS region and enter data volume, object count, wave size, archive mix, restore tier, retention and FastConnect/Direct Connect assumptions."),
        ("3. AWS estimate", "Shows charged AWS-side components. Outbound is optional and can use a specific FastConnect/Direct Connect rate when entered."),
        ("4. OCI estimate", "Separates one-time write requests, storage accrued while data arrives, and the recurring final Object Storage cost."),
        ("5. Early delete", "Add one or more groups of Deep Archive objects by volume and average archive age. The model calculates the 180-day minimum-storage remainder."),
        ("6. Summary", "Use this tab for the executive view. It intentionally keeps recurring OCI storage separate from one-time migration costs."),
        ("Important", "All TB/GB inputs use decimal units (1 TB = 1,000 GB), matching cloud price-list units. Taxes, credits, enterprise discounts, source-storage charges, VM cost and partner circuit charges are excluded unless you add them manually."),
    ]
    table(ws, 4, ["Step", "How to use"], [28, 110])
    for r, (name, text) in enumerate(items, 5):
        ws.cell(r, 1, name); ws.cell(r, 2, text)
        body(ws.cell(r, 1)); body(ws.cell(r, 2))
        ws.row_dimensions[r].height = 44
    ws.freeze_panes = "A4"


def build_rates(wb):
    ws = wb.create_sheet("Tariffs by region")
    title(ws, "Tariffs by region", "Enter public list prices or contract rates. A blank rate causes the associated calculation to show ‘Not priced’ rather than assume zero.")
    headers = [
        "AWS region", "Currency", "AWS outbound\nUSD/GB", "FastConnect / Direct Connect\noutbound USD/GB", "Deep Archive storage\nUSD/GB-month", "Deep Archive BULK retrieval\nUSD/GB", "Deep Archive STANDARD retrieval\nUSD/GB", "Temporary S3 Standard restore\nUSD/GB-month", "S3 PUT/LIST\nUSD/1,000", "S3 GET/TAG\nUSD/1,000", "Batch job\nUSD/job", "Batch object\nUSD/1,000", "OCI Standard storage\nUSD/GB-month", "OCI writes\nUSD/10,000", "Notes / source"
    ]
    widths = [16, 11, 15, 24, 20, 23, 27, 26, 17, 17, 15, 20, 22, 17, 42]
    table(ws, 4, headers, widths)
    # Current values collected in Raijin. Deep Archive storage stays deliberately
    # blank: it must be confirmed from the customer's applicable price list before
    # an early-deletion decision is made.
    starter = [
        ["sa-east-1", "USD", 0.15, None, None, 0.008, 0.028, 0.0405, 0.007, 0.00056, 0.25, 0.000015, 0.0255, 0.0034, "Raijin public-price collection 2026-08-19; confirm contract and Deep Archive storage price."],
        ["us-east-1", "USD", 0.09, None, None, 0.0025, 0.02, 0.0125, 0.055, 0.0004, 0.25, 0.001, 0.0255, 0.0034, "Example only; refresh with the applicable public or contracted rate."],
    ]
    for r, values in enumerate(starter, 5):
        for c, value in enumerate(values, 1):
            cell = ws.cell(r, c, value)
            body(cell, input_cell=True)
            if 3 <= c <= 14:
                currency(cell)
    for r in range(7, 55):
        for c in range(1, 16):
            body(ws.cell(r, c), input_cell=True)
            if 3 <= c <= 14:
                currency(ws.cell(r, c))
    ws.auto_filter.ref = "A4:O54"
    ws.freeze_panes = "A5"


def build_inputs(wb):
    ws = wb.create_sheet("Main inputs")
    title(ws, "Main inputs", "Yellow fields are intentionally customizable. Select the region first; rates are retrieved from ‘Tariffs by region’. Boolean fields are TRUE/FALSE.")
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 78
    table(ws, 4, ["Parameter", "Value", "Description"], [38, 24, 78])
    values = [
        ("Scenario name", "600 TB Deep Archive migration", "Label used only in the workbook."),
        ("AWS region", "sa-east-1", "Must match a row in ‘Tariffs by region’ exactly."),
        ("Total data (TB)", 600, "Decimal TB; 1 TB = 1,000 GB."),
        ("Total objects", 3600000, "Total number of S3 objects in the source."),
        ("Wave size (TB)", 5, "Maximum data per wave. The model rounds up the wave count."),
        ("Deep Archive share", 1, "Share of total data in S3 DEEP_ARCHIVE. Use 1 for 100%."),
        ("Glacier Flexible share", 0, "Share of total data in S3 GLACIER / Flexible Retrieval."),
        ("Non-archive share", 0, "Share of total data already readable without restore."),
        ("Restore tier", "BULK", "BULK or STANDARD. Applies to Deep Archive and Glacier Flexible data."),
        ("Temporary restore retention (hours)", 48, "Retention starts after restore becomes available. 48h is the safer Deep Archive BULK default."),
        ("Expected restore poll cycles / wave", 24, "Adjust to the polling policy and expected restore wait."),
        ("Preserve S3 tags", True, "Adds one S3 GET/TAG request per object."),
        ("Include AWS outbound", True, "Turn off when outbound is billed externally or intentionally excluded."),
        ("Use FastConnect / Direct Connect outbound rate", True, "Uses the optional regional circuit/egress rate when it is populated; otherwise uses standard AWS outbound."),
        ("Include OCI storage", True, "Includes OCI write requests, migration-period storage accrual and recurring final storage."),
        ("Estimated migration duration (days)", 47, "Used only to approximate OCI storage accrued while data lands linearly."),
        ("OCI write operations per object", 4, "Average final write operations per object. Use 1 for small PutObject files; increase for multipart."),
        ("Include early Deep Archive delete", False, "Adds the ‘Early delete’ worksheet total to the one-time estimate."),
    ]
    for r, (label, value, description) in enumerate(values, 5):
        ws.cell(r, 1, label); ws.cell(r, 2, value); ws.cell(r, 3, description)
        body(ws.cell(r, 1)); body(ws.cell(r, 2), input_cell=True); body(ws.cell(r, 3))
        ws.row_dimensions[r].height = 30
    # Calculated fields
    calculated_at = 25
    table(ws, calculated_at, ["Calculated planning metric", "Value", "Explanation"], [38, 24, 78])
    calculated = [
        ("Number of waves", "=ROUNDUP(B7/B9,0)", "Rounded up; final wave can be smaller."),
        ("Average objects / wave", "=ROUNDUP(B8/B26,0)", "Useful for Batch, polling and request estimates."),
        ("Source ListObjectsV2 pages", "=ROUNDUP(B8/1000,0)", "One discovery page is up to 1,000 objects."),
        ("Archive mix validation", "=B10+B11+B12", "Must equal 1.00; red means the shares are inconsistent."),
        ("Effective outbound rate USD/GB", "=IF(B18,IFERROR(IF(VLOOKUP(B6,'Tariffs by region'!A:O,4,FALSE)>0,VLOOKUP(B6,'Tariffs by region'!A:O,4,FALSE),VLOOKUP(B6,'Tariffs by region'!A:O,3,FALSE)),VLOOKUP(B6,'Tariffs by region'!A:O,3,FALSE)),VLOOKUP(B6,'Tariffs by region'!A:O,3,FALSE))", "FastConnect/Direct Connect rate wins only when enabled and entered in the tariff table."),
    ]
    for r, (label, formula, description) in enumerate(calculated, calculated_at + 1):
        ws.cell(r, 1, label); ws.cell(r, 2, formula); ws.cell(r, 3, description)
        body(ws.cell(r, 1)); body(ws.cell(r, 2)); body(ws.cell(r, 3))
        if r != 28:
            number(ws.cell(r, 2), 2)
    ws["B29"].fill = PatternFill("solid", fgColor=YELLOW)
    ws["B29"].font = Font(bold=True, color="111827")
    ws["B30"].fill = PatternFill("solid", fgColor=TEAL)
    ws["B30"].font = Font(bold=True, color="111827")
    currency(ws["B30"])
    ws.conditional_formatting.add("B29", ColorScaleRule(start_type="num", start_value=0, start_color="F8696B", mid_type="num", mid_value=1, mid_color="FFEB84", end_type="num", end_value=1, end_color="63BE7B"))
    truth = DataValidation(type="list", formula1='"TRUE,FALSE"')
    tier = DataValidation(type="list", formula1='"BULK,STANDARD"')
    ws.add_data_validation(truth); truth.add("B16:B19"); truth.add("B22")
    ws.add_data_validation(tier); tier.add("B13")
    ws.freeze_panes = "A5"


def build_aws(wb):
    ws = wb.create_sheet("AWS estimate")
    title(ws, "AWS one-time migration estimate", "Uses the selected regional rate. Polling is intentionally conservative: it charges a ListObjectsV2-style cycle for each wave. You may override any tariff in the regional table.")
    headers = ["Component", "Quantity", "Unit", "Rate source", "Rate", "Estimated cost", "Notes"]
    table(ws, 4, headers, [40, 20, 18, 18, 18, 20, 58])
    # label, quantity formula, unit, rate formula, enabled formula, note
    rows = [
        ("Deep Archive BULK retrieval", "='Main inputs'!B9*1000*'Main inputs'!B10", "GB", "=VLOOKUP('Main inputs'!B6,'Tariffs by region'!A:O,6,FALSE)", "='Main inputs'!B13=\"BULK\"", "Only data in Deep Archive."),
        ("Deep Archive STANDARD retrieval", "='Main inputs'!B9*1000*'Main inputs'!B10", "GB", "=VLOOKUP('Main inputs'!B6,'Tariffs by region'!A:O,7,FALSE)", "='Main inputs'!B13=\"STANDARD\"", "Only data in Deep Archive."),
        ("Temporary restored S3 Standard copy", "='Main inputs'!B9*1000*('Main inputs'!B10+'Main inputs'!B11)*'Main inputs'!B14/24/30", "GB-month", "=VLOOKUP('Main inputs'!B6,'Tariffs by region'!A:O,8,FALSE)", "=TRUE", "Retention is charged only while the restored copy exists."),
        ("S3 Batch Operations jobs", "=1", "job / wave", "=VLOOKUP('Main inputs'!B6,'Tariffs by region'!A:O,11,FALSE)", "=TRUE", "One Batch restore job per archive wave."),
        ("S3 Batch Operations object tasks", "='Main inputs'!B27", "objects / wave", "=VLOOKUP('Main inputs'!B6,'Tariffs by region'!A:O,12,FALSE)/1000", "=TRUE", "Rate is normalized to USD/object."),
        ("Allocated discovery ListObjectsV2", "=ROUNDUP('Main inputs'!B27/1000,0)", "requests / wave", "=VLOOKUP('Main inputs'!B6,'Tariffs by region'!A:O,9,FALSE)/1000", "=TRUE", "Discovery itself runs once; allocation is proportional to objects in the wave."),
        ("Restore polling — conservative", "='Main inputs'!B28*'Main inputs'!B15", "requests / wave", "=VLOOKUP('Main inputs'!B6,'Tariffs by region'!A:O,9,FALSE)/1000", "=TRUE", "Conservative ListObjectsV2 equivalent. Actual Raijin adaptive polling can be lower."),
        ("S3 object reads", "='Main inputs'!B27", "requests / wave", "=VLOOKUP('Main inputs'!B6,'Tariffs by region'!A:O,10,FALSE)/1000", "=TRUE", "One streaming GetObject per object."),
        ("S3 object-tag reads", "='Main inputs'!B27", "requests / wave", "=VLOOKUP('Main inputs'!B6,'Tariffs by region'!A:O,10,FALSE)/1000", "='Main inputs'!B16", "Only when tag preservation is enabled."),
        ("AWS transfer out to OCI", "='Main inputs'!B9*1000", "GB", "='Main inputs'!B30", "='Main inputs'!B17", "Uses FastConnect/Direct Connect rate only if selected and populated."),
    ]
    for r, row in enumerate(rows, 5):
        label, qty, unit, rate, enabled, note = row
        values = [label, qty, unit, "Contract/public tariff", rate, f'=IF({enabled},IFERROR(B{r}*E{r},\"Not priced\"),0)', note]
        for c, value in enumerate(values, 1):
            ws.cell(r, c, value); body(ws.cell(r, c))
        number(ws.cell(r, 2), 4); currency(ws.cell(r, 5)); currency(ws.cell(r, 6))
    total_row = 16
    ws.cell(total_row, 1, "AWS total / wave"); ws.cell(total_row, 6, "=SUM(F5:F14)")
    ws.cell(total_row + 1, 1, "AWS total / project"); ws.cell(total_row + 1, 6, "=F16*'Main inputs'!B26")
    for r in (total_row, total_row + 1):
        for c in range(1, 7): body(ws.cell(r, c))
        ws.cell(r, 1).font = Font(bold=True, color=WHITE); ws.cell(r, 6).font = Font(size=14, bold=True, color=WHITE); currency(ws.cell(r, 6))
        ws.cell(r, 6).fill = PatternFill("solid", fgColor=GREEN)
    ws.freeze_panes = "A5"


def build_oci(wb):
    ws = wb.create_sheet("OCI estimate")
    title(ws, "OCI cost estimate", "OCI storage is shown separately because it becomes a recurring destination cost. Migration-period storage assumes data arrives evenly over the configured duration.")
    table(ws, 4, ["Component", "Quantity", "Unit", "Rate", "Estimated cost", "Notes"], [44, 24, 18, 18, 22, 62])
    rows = [
        ("OCI Object Storage write requests / wave", "='Main inputs'!B27*'Main inputs'!B21", "operations", "=VLOOKUP('Main inputs'!B6,'Tariffs by region'!A:O,14,FALSE)/10000", "='Main inputs'!B19", "Average operations/object is customizable for multipart."),
        ("OCI final storage / wave / month", "='Main inputs'!B9*1000", "GB-month", "=VLOOKUP('Main inputs'!B6,'Tariffs by region'!A:O,13,FALSE)", "='Main inputs'!B19", "Recurring once this wave is fully present."),
        ("OCI storage accrued during full migration", "='Main inputs'!B7*1000*'Main inputs'!B20/2/30", "GB-month", "=VLOOKUP('Main inputs'!B6,'Tariffs by region'!A:O,13,FALSE)", "='Main inputs'!B19", "Linear-arrival approximation; excludes later recurring storage."),
        ("OCI final storage / project / month", "='Main inputs'!B7*1000", "GB-month", "=VLOOKUP('Main inputs'!B6,'Tariffs by region'!A:O,13,FALSE)", "='Main inputs'!B19", "Recurring final destination storage."),
    ]
    for r, (label, qty, unit, rate, enabled, note) in enumerate(rows, 5):
        vals = [label, qty, unit, rate, f'=IF({enabled},IFERROR(B{r}*D{r},\"Not priced\"),0)', note]
        for c, v in enumerate(vals, 1): ws.cell(r, c, v); body(ws.cell(r, c))
        number(ws.cell(r, 2), 4); currency(ws.cell(r, 4)); currency(ws.cell(r, 5))
    ws.cell(11, 1, "OCI one-time migration cost"); ws.cell(11, 5, "=E5+E7")
    ws.cell(12, 1, "OCI recurring final monthly cost"); ws.cell(12, 5, "=E8")
    for r in (11, 12):
        for c in range(1, 6): body(ws.cell(r, c))
        ws.cell(r, 1).font = Font(bold=True, color=WHITE); ws.cell(r, 5).font = Font(size=14, bold=True, color=WHITE); ws.cell(r, 5).fill = PatternFill("solid", fgColor=GREEN); currency(ws.cell(r, 5))


def build_early_delete(wb):
    ws = wb.create_sheet("Early Deep Archive delete")
    title(ws, "Deep Archive early-delete model", "Use one row per age group. AWS Deep Archive has a 180-day minimum storage duration; this model estimates the pro-rated remaining storage charge. Enter the verified regional Deep Archive storage rate before using this result.")
    table(ws, 4, ["Group / prefix", "Deep Archive data to delete (TB)", "Average age in Deep Archive (days)", "Deep Archive rate USD/GB-month", "Remaining minimum days", "Estimated early-delete charge", "Notes"], [30, 28, 32, 31, 24, 30, 55])
    for r in range(5, 25):
        values = [f"Group {r - 4}", None, None, "=VLOOKUP('Main inputs'!B6,'Tariffs by region'!A:O,5,FALSE)", f'=IF(OR(B{r}=\"\",C{r}=\"\"),\"\",MAX(0,180-C{r}))', f'=IF(OR(B{r}=\"\",C{r}=\"\",D{r}=\"\"),\"Not priced\",B{r}*1000*D{r}*E{r}/30)', None]
        for c, v in enumerate(values, 1):
            ws.cell(r, c, v); body(ws.cell(r, c), input_cell=c in (1, 2, 3, 7))
        number(ws.cell(r, 2), 4); number(ws.cell(r, 3), 1); currency(ws.cell(r, 4)); number(ws.cell(r, 5), 1); currency(ws.cell(r, 6))
    ws.cell(26, 1, "Total early-delete charge"); ws.cell(26, 6, "=SUM(F5:F24)")
    for c in range(1, 7): body(ws.cell(26, c))
    ws.cell(26, 1).font = Font(bold=True, color=WHITE); ws.cell(26, 6).fill = PatternFill("solid", fgColor=ORANGE); ws.cell(26, 6).font = Font(size=14, bold=True, color=WHITE); currency(ws.cell(26, 6))
    ws.freeze_panes = "A5"


def build_summary(wb):
    ws = wb.create_sheet("Summary")
    title(ws, "Executive summary", "One-time AWS migration, optional early delete and OCI are separated to avoid confusing migration cost with the recurring destination-storage expense.")
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 27
    ws.column_dimensions["C"].width = 84
    table(ws, 4, ["Metric", "Estimated value", "Interpretation"], [52, 27, 84])
    rows = [
        ("Scenario", "='Main inputs'!B5", "Editable scenario label."),
        ("Data volume", "='Main inputs'!B7", "Decimal TB."),
        ("Number of waves", "='Main inputs'!B26", "Wave size is configurable."),
        ("AWS one-time cost / wave", "='AWS estimate'!F16", "Includes retrieval, temporary restore, requests and optional outbound."),
        ("AWS one-time cost / project", "='AWS estimate'!F17", "Public or contracted rate depending on the tariff table."),
        ("Early Deep Archive delete", "=IF('Main inputs'!B22,'Early Deep Archive delete'!F26,0)", "Only included when explicitly enabled."),
        ("OCI one-time migration cost", "='OCI estimate'!E11", "Writes plus storage accrued while the migration is running."),
        ("Migration one-time total", "=B9+B10+B11", "Excludes future recurring OCI storage."),
        ("OCI recurring final storage / month", "='OCI estimate'!E12", "Expected monthly cost after all data is present in OCI."),
    ]
    for r, (label, formula, note) in enumerate(rows, 5):
        ws.cell(r, 1, label); ws.cell(r, 2, formula); ws.cell(r, 3, note)
        for c in range(1, 4): body(ws.cell(r, c))
        if r not in (5, 6, 7): currency(ws.cell(r, 2))
    number(ws["B6"], 2); number(ws["B7"], 0)
    for r in (12, 13):
        ws.cell(r, 2).fill = PatternFill("solid", fgColor=GREEN)
        ws.cell(r, 2).font = Font(size=14, bold=True, color=WHITE)
    ws["A16"] = "Decision reminders"
    ws["A16"].font = Font(size=14, bold=True, color=WHITE); ws["A16"].fill = PatternFill("solid", fgColor=BLUE)
    ws.merge_cells("A17:C21")
    ws["A17"] = ("• For Deep Archive BULK, 48 hours is the safer temporary-copy retention default.\n"
                  "• FastConnect/Direct Connect can change the outbound cost, but partner circuit charges are not modeled unless entered manually.\n"
                  "• 600 TB is above AWS’s 500 TB/month public-price threshold: obtain a written commercial quote and enter the agreed outbound rate.\n"
                  "• Early-delete values require the true Deep Archive age distribution; last-modified date alone may not be the transition date.")
    body(ws["A17"]); ws["A17"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[17].height = 92
    ws.freeze_panes = "A5"


def main():
    wb = Workbook()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    build_readme(wb)
    build_rates(wb)
    build_inputs(wb)
    build_aws(wb)
    build_oci(wb)
    build_early_delete(wb)
    build_summary(wb)
    for ws in wb.worksheets:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
