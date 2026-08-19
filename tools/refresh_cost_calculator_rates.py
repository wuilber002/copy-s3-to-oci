#!/usr/bin/env python3
"""Refresh the workbook's AWS public tariffs from Raijin's public endpoints.

No AWS credentials are used.  This intentionally mirrors the narrow AWS price
mapping in app.main, so the workbook and Raijin consume the same two regional
catalogs and units.
"""

from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path
from urllib.request import Request, urlopen

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path(__file__).resolve().parents[1] / "docs" / "raijin-cost-calculator.xlsx"
S3_URL = "https://pricing.us-east-1.amazonaws.com/offers/v1.0/aws/AmazonS3/current/{region}/index.json"
TRANSFER_URL = "https://pricing.us-east-1.amazonaws.com/offers/v1.0/aws/AWSDataTransfer/current/{region}/index.json"
SUPPORTED_REGIONS = {"sa-east-1", "us-east-1"}


def first_on_demand_usd(terms: dict) -> float | None:
    # Keep traversal equivalent to Raijin, but start at the selected SKU term.
    def walk(value):
        if not isinstance(value, dict):
            return None
        price = (value.get("pricePerUnit") or {}).get("USD")
        if price is not None:
            try:
                return float(price)
            except (TypeError, ValueError):
                return None
        for child in value.values():
            found = walk(child)
            if found is not None:
                return found
        return None

    return walk(terms.get("OnDemand") or {})


def fetch_json(url: str) -> dict:
    request = Request(url, headers={"User-Agent": "raijin-cost-calculator/1.0"})
    with urlopen(request, timeout=60) as response:
        body = response.read(64 * 1024 * 1024 + 1)
    if len(body) > 64 * 1024 * 1024:
        raise RuntimeError("AWS public pricing catalog exceeds the 64 MiB safety limit")
    return json.loads(body.decode("utf-8"))


def s3_rates(catalog: dict) -> dict[str, float]:
    rates: dict[str, float] = {}
    for product in (catalog.get("products") or {}).values():
        attrs = product.get("attributes") or {}
        blob = " ".join(str(attrs.get(key, "")).lower() for key in (
            "group", "groupDescription", "feeCode", "feeDescription", "usagetype", "operation",
            "storageClass", "volumeType", "productFamily",
        )).replace("-", " ").replace("_", " ").replace("/", " ")
        sku = product.get("sku")
        price = first_on_demand_usd({"OnDemand": (catalog.get("terms") or {}).get("OnDemand", {}).get(sku, {})}) if sku else None
        if price is None:
            continue
        gib_price = price * (1024 ** 3 / 1_000_000_000)
        if "batch" in blob and "job" in blob:
            rates.setdefault("batch_job", price)
        elif "batch" in blob and ("object" in blob or "operation" in blob):
            rates.setdefault("batch_object_per_1000", price * 1000)
        elif "tier1" in blob or "put/copy/post/list" in blob:
            rates.setdefault("put_list_per_1000", price * 1000)
        elif "tier2" in blob or "get and all other" in blob:
            rates.setdefault("get_tag_per_1000", price * 1000)
        elif "deep" in blob and "retrieval" in blob and "bulk" in blob:
            rates.setdefault("deep_bulk_per_gib", gib_price)
        elif "deep" in blob and "retrieval" in blob and "standard" in blob:
            rates.setdefault("deep_standard_per_gib", gib_price)
        elif "timedstorage" in blob and "standard" in blob and "archive" not in blob:
            rates.setdefault("temporary_standard_per_gib_month", gib_price)
    return rates


def transfer_rate(catalog: dict) -> float | None:
    terms = (catalog.get("terms") or {}).get("OnDemand") or {}
    for product in (catalog.get("products") or {}).values():
        attrs = product.get("attributes") or {}
        if attrs.get("toLocation") != "External" or attrs.get("transferType") != "AWS Outbound" or attrs.get("fromLocation") in {None, "", "Global"}:
            continue
        price = first_on_demand_usd({"OnDemand": terms.get(product.get("sku"), {})})
        if price is not None:
            return price * (1024 ** 3 / 1_000_000_000)
    return None


def update_workbook(workbook: Path, regions: list[str]) -> None:
    wb = load_workbook(workbook)
    ws = wb["Tariffs by region"]
    rows = {str(ws.cell(row, 1).value): row for row in range(5, ws.max_row + 1) if ws.cell(row, 1).value}
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    for region in regions:
        rates = s3_rates(fetch_json(S3_URL.format(region=region)))
        outbound = transfer_rate(fetch_json(TRANSFER_URL.format(region=region)))
        if outbound is not None:
            rates["outbound_per_gib"] = outbound
        row = rows.get(region)
        if not row:
            row = max(5, ws.max_row + 1)
            ws.cell(row, 1, region)
            ws.cell(row, 2, "USD")
        # Workbook labels are decimal GB. Raijin converts its native catalog
        # values to GiB, so convert back here for a familiar spreadsheet unit.
        gib_to_gb = 1_000_000_000 / (1024 ** 3)
        mapping = {
            3: ("outbound_per_gib", gib_to_gb), 5: ("deep_bulk_per_gib", gib_to_gb),
            6: ("deep_standard_per_gib", gib_to_gb), 7: ("temporary_standard_per_gib_month", gib_to_gb),
            8: ("put_list_per_1000", 1), 9: ("get_tag_per_1000", 1),
            10: ("batch_job", 1), 11: ("batch_object_per_1000", 1),
        }
        for column, (key, multiplier) in mapping.items():
            if key in rates:
                ws.cell(row, column, rates[key] * multiplier)
        ws.cell(row, 12, f"AWS public catalogs refreshed {now}. Deep Archive storage remains a separate public-price input.")
        print(f"{region}: refreshed {len(rates)} AWS rate(s) into row {row}")
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(workbook)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--regions", default="sa-east-1,us-east-1", help="Comma-separated supported regions: sa-east-1, us-east-1")
    args = parser.parse_args()
    regions = [region.strip() for region in args.regions.split(",") if region.strip()]
    invalid = sorted(set(regions) - SUPPORTED_REGIONS)
    if invalid:
        parser.error(f"Unsupported region(s): {', '.join(invalid)}. Supported: {', '.join(sorted(SUPPORTED_REGIONS))}")
    if not args.workbook.is_file():
        parser.error(f"Workbook does not exist: {args.workbook}")
    update_workbook(args.workbook, regions)


if __name__ == "__main__":
    main()
